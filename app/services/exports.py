"""Shared CSV/Excel/PDF export helpers so every report offers all three
formats from the same (header, rows) shape instead of hand-rolling one-off
export code per report."""

import csv
import io

from flask import Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def _stringify_row(row):
    return [("" if cell is None else str(cell)) for cell in row]


def csv_response(filename, header, rows):
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(header)
    writer.writerows(_stringify_row(r) for r in rows)
    return Response(
        buf.getvalue(), mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def excel_response(filename, header, rows, sheet_title="Report"):
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_title or "Report")[:31]
    ws.append(list(header))
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for row in rows:
        ws.append(_stringify_row(row))
    for i in range(1, len(header) + 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def pdf_response(filename, title, header, rows, subtitle=None):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(letter), title=title,
        topMargin=0.5 * inch, bottomMargin=0.5 * inch, leftMargin=0.5 * inch, rightMargin=0.5 * inch,
    )
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles["Title"])]
    if subtitle:
        elements.append(Paragraph(subtitle, styles["Normal"]))
    elements.append(Spacer(1, 12))

    table_data = [list(header)] + [_stringify_row(r) for r in rows]
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4f46e5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    return Response(
        buf.getvalue(), mimetype="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def export_response(fmt, base_filename, header, rows, title=None, subtitle=None):
    """Dispatches to the right format. `fmt` is the `export` query param
    value: 'csv' | 'xlsx' | 'pdf'."""
    if fmt == "xlsx":
        return excel_response(f"{base_filename}.xlsx", header, rows, sheet_title=title or base_filename)
    if fmt == "pdf":
        return pdf_response(f"{base_filename}.pdf", title or base_filename, header, rows, subtitle=subtitle)
    return csv_response(f"{base_filename}.csv", header, rows)
